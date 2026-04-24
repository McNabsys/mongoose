"""Unit tests for the Excel manifest parser + yellow-fill detection.

Uses synthetic openpyxl workbooks so tests run without the real file.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from mongoose.etl.excel_manifest import (
    CONC_TO_GROUP,
    EXPECTED_BIOCHEM_FLAGGED,
    HEADER_ROW,
    SHEET_NAME,
    _is_yellow,
    load_excel_manifest,
)


YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
WHITE = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


def _write_minimal_workbook(path, runs, headers=("Run ID", "Date", "Purpose")):
    """runs: list of (run_id, conc_raw, is_yellow). Writes a fake Excel with
    headers at row 4 and runs at rows 5..."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # Row 4: headers. Fill columns up to index 10 so Conc (col 10) is set.
    for col, name in enumerate(headers, start=1):
        ws.cell(HEADER_ROW, col, name)
    ws.cell(HEADER_ROW, 10, "Conc")
    for i, (rid, conc, yellow) in enumerate(runs):
        row = HEADER_ROW + 1 + i
        ws.cell(row, 1, rid)
        ws.cell(row, 2, "2025-01-01")
        ws.cell(row, 3, "test")
        ws.cell(row, 10, conc)
        if yellow:
            ws.cell(row, 1).fill = YELLOW
        else:
            ws.cell(row, 1).fill = WHITE
    wb.save(path)


def test_is_yellow_classifier():
    assert _is_yellow("FFFFFF00") is True
    assert _is_yellow("00FFFF00") is True  # transparent yellow still counts
    assert _is_yellow("FFFFFFCC") is True
    assert _is_yellow("FFFFFFFF") is False
    assert _is_yellow("FF00FF00") is False  # green
    assert _is_yellow("") is False
    assert _is_yellow(None) is False


def test_parses_canonical(tmp_path):
    p = tmp_path / "m.xlsx"
    runs = [(rid, "10 ng/uL", True) for rid in sorted(EXPECTED_BIOCHEM_FLAGGED)]
    runs += [("STB03-XXX-00", "10 ng/uL", False)]
    _write_minimal_workbook(p, runs)
    parsed = load_excel_manifest(p)
    assert len(parsed) == len(runs)
    flagged_ids = {r.run_id for r in parsed if r.biochem_flagged_good}
    assert flagged_ids == EXPECTED_BIOCHEM_FLAGGED


def test_biochem_mismatch_raises(tmp_path):
    p = tmp_path / "m.xlsx"
    # Make a DIFFERENT set yellow -- validation should reject.
    runs = [
        ("STB03-UNKNOWN-AAA", "10 ng/uL", True),  # not in canonical set
    ]
    _write_minimal_workbook(p, runs)
    with pytest.raises(ValueError, match="Yellow-fill detection mismatch"):
        load_excel_manifest(p)


def test_biochem_validation_can_be_bypassed(tmp_path):
    p = tmp_path / "m.xlsx"
    runs = [("STB03-NEW-RUN-99", "10 ng/uL", True)]
    _write_minimal_workbook(p, runs)
    parsed = load_excel_manifest(p, validate_biochem=False)
    assert len(parsed) == 1
    assert parsed[0].biochem_flagged_good is True


def test_conc_to_group_mapping_complete():
    """Spec §39: 3 concentration groups, canonical names std/low/low_dil."""
    assert set(CONC_TO_GROUP.values()) == {"std", "low", "low_dil"}


def test_unknown_conc_raises(tmp_path):
    p = tmp_path / "m.xlsx"
    runs = [("STB03-X-00", "17 mg/ml", False)]
    _write_minimal_workbook(p, runs)
    with pytest.raises(ValueError, match="Unknown Conc"):
        load_excel_manifest(p, validate_biochem=False)
