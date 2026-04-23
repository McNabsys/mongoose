"""Parser for the Nabsys project-level Excel manifest.

File: ``Project Mongoose - input data.xlsx``, sheet ``Short runs``.

Headers live on row 4; data rows start at row 5. Each data row is one
run. The ``Run ID`` column (A) matches the on-disk directory name under
``E. coli/{Black,Blue,Red}/<run_id>/``.

Additionally, the biochem team flags 6 of 30 runs as "good" by painting
column A yellow. We detect this by reading ``cell.fill.start_color.rgb``
and matching ``FF(FFFF00|FFFFCC|...)``. The expected flagged list is
validated against a spec-provided canonical set; a mismatch stops the
ETL so a silent change in the highlighting can't corrupt downstream
stratification.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ROW = 4
SHEET_NAME = "Short runs"
RUN_ID_COLUMN = 1


# Map from raw Excel ``Conc`` string to the canonical concentration_group
# categorical. The color folders on disk (E. coli/Black, Red, Blue) mirror
# this mapping 1:1 — verified 2026-04-23 via cross-reference against the
# Run IDs present in each folder.
CONC_TO_GROUP: dict[str, str] = {
    "10 ng/uL": "std",       # Black folder
    "5 ng/uL": "low",        # Red folder
    "5 ng/uL, diluted": "low_dil",  # Blue folder
}

# Canonical biochem-flagged set from the phase 0 spec (§51-59). The ETL
# asserts the yellow-detected set matches this exactly so a silent change
# in the highlighting cannot silently corrupt every downstream
# stratification. Spec: "If the openpyxl-detected set doesn't exactly
# match these 6 IDs, stop and surface the discrepancy to the user."
EXPECTED_BIOCHEM_FLAGGED: frozenset[str] = frozenset({
    "STB03-060C-02L58270w05-433H09a",
    "STB03-060A-02L58270w05-433B23e",
    "STB03-064D-02L58270w05-433H09d",
    "STB03-064B-02L58270w05-202G16g",
    "STB03-065F-02L58270w05-433H09h",
    "STB03-065A-02L58270w05-433B23f",
})


@dataclass(frozen=True)
class RunMetadata:
    """Per-run metadata extracted from the Excel sheet, untouched types."""

    run_id: str
    row: int
    conc_raw: str
    concentration_group: str  # one of: std, low, low_dil
    biochem_flagged_good: bool
    fields: dict[str, Any]  # header name → cell value, covers all populated columns


def _is_yellow(rgb: Any) -> bool:
    """Yellow-fill detection for biochem-flagged rows.

    Excel stores fills as ARGB strings (e.g. ``FFFFFF00`` = opaque yellow).
    We tolerate the common yellow variants Excel may emit — the first two
    bytes are alpha, the remaining six are RGB.
    """
    if not rgb:
        return False
    s = str(rgb).upper()
    if len(s) < 6:
        return False
    # Strip ARGB alpha if present, compare RGB hex.
    rgb_hex = s[-6:]
    # Pure Excel yellow is FFFF00; tolerate FFFFCC (light yellow) as well.
    return rgb_hex in {"FFFF00", "FFFFCC", "FFFF99", "FFFFE0"}


def _normalize_conc(conc_raw: str) -> str:
    """Normalize the raw Conc string to the canonical group."""
    if conc_raw not in CONC_TO_GROUP:
        raise ValueError(
            f"Unknown Conc value {conc_raw!r}; expected one of "
            f"{sorted(CONC_TO_GROUP.keys())}"
        )
    return CONC_TO_GROUP[conc_raw]


def load_excel_manifest(
    path: Path | str,
    *,
    sheet: str = SHEET_NAME,
    validate_biochem: bool = True,
) -> list[RunMetadata]:
    """Parse the Short-runs sheet into a list of RunMetadata.

    Args:
        path: Path to ``Project Mongoose - input data.xlsx``.
        sheet: Sheet name to read. Default ``"Short runs"``.
        validate_biochem: If True (default), assert that the yellow-
            highlighted Run IDs exactly match EXPECTED_BIOCHEM_FLAGGED.
            Raises ``ValueError`` on mismatch. Set False for unit tests
            or if the canonical list is deliberately being updated.

    Returns:
        Runs in source-sheet order.
    """
    path = Path(path)
    # data_only=False so ``cell.fill`` carries its format; we still read
    # computed values (string Conc, numeric metrics) just fine this way.
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"{path}: sheet {sheet!r} not found; sheets={wb.sheetnames}")
    ws = wb[sheet]

    # Header row (row 4). Collect header name → column index map; skip Nones.
    headers: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(HEADER_ROW, col).value
        if name is not None:
            headers[col] = str(name).strip()

    if RUN_ID_COLUMN not in headers or headers[RUN_ID_COLUMN] != "Run ID":
        raise ValueError(
            f"{path}: expected 'Run ID' at col {RUN_ID_COLUMN} row {HEADER_ROW}, "
            f"got {headers.get(RUN_ID_COLUMN)!r}"
        )

    # Reload WITHOUT data_only so fills are preserved. We already have values
    # from the first pass; this pass only supplies cell-fill colors.
    wb_fmt = load_workbook(path)
    ws_fmt = wb_fmt[sheet]

    runs: list[RunMetadata] = []
    flagged: set[str] = set()

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        run_id = ws.cell(row, RUN_ID_COLUMN).value
        if not run_id or not isinstance(run_id, str) or not run_id.startswith("STB"):
            continue

        conc_raw = ws.cell(row, 10).value
        if conc_raw is None:
            raise ValueError(f"{path} row {row}: missing Conc for run {run_id}")
        conc_raw = str(conc_raw).strip()
        concentration_group = _normalize_conc(conc_raw)

        # Detect yellow fill on the Run ID cell (col A). Matches spec
        # §48-50: "cell.fill.start_color.rgb ... any row with a yellow
        # fill ... gets biochem_flagged_good = True."
        fill_color = ws_fmt.cell(row, RUN_ID_COLUMN).fill.start_color
        biochem = _is_yellow(fill_color.rgb) if fill_color else False
        if biochem:
            flagged.add(run_id)

        fields: dict[str, Any] = {}
        for col, name in headers.items():
            fields[name] = ws.cell(row, col).value

        runs.append(
            RunMetadata(
                run_id=run_id,
                row=row,
                conc_raw=conc_raw,
                concentration_group=concentration_group,
                biochem_flagged_good=biochem,
                fields=fields,
            )
        )

    if validate_biochem and flagged != EXPECTED_BIOCHEM_FLAGGED:
        missing = EXPECTED_BIOCHEM_FLAGGED - flagged
        extra = flagged - EXPECTED_BIOCHEM_FLAGGED
        raise ValueError(
            f"Yellow-fill detection mismatch in {path}:\n"
            f"  missing (expected yellow, not detected): {sorted(missing)}\n"
            f"  extra (detected yellow, not expected): {sorted(extra)}\n"
            f"Either the Excel highlighting changed, the color detection is "
            f"wrong, or the canonical list in excel_manifest.py is stale. "
            f"Do not proceed until this is resolved manually."
        )

    return runs
